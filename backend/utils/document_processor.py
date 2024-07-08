import os
from docx import Document
from openpyxl import load_workbook
import PyPDF2
from pdf2image import convert_from_path
import pytesseract
from PIL import Image

def process_document(file_path):
    _, file_extension = os.path.splitext(file_path)
    file_extension = file_extension.lower()

    if file_extension == '.txt':
        return process_text(file_path)
    elif file_extension == '.docx':
        return process_docx(file_path)
    elif file_extension in ['.xls', '.xlsx']:
        return process_excel(file_path)
    elif file_extension == '.pdf':
        return process_pdf(file_path)
    elif file_extension in ['.png', '.jpg', '.jpeg']:
        return process_image(file_path)
    else:
        print(f"Unsupported file type: {file_extension}")
        return None, None

def process_text(file_path):
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()
    return content, {"source": file_path}

def process_docx(file_path):
    doc = Document(file_path)
    content = "\n".join([paragraph.text for paragraph in doc.paragraphs])
    return content, {"source": file_path}

def process_excel(file_path):
    wb = load_workbook(file_path)
    content = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows(values_only=True):
            content.append(" ".join(str(cell) for cell in row if cell))
    return "\n".join(content), {"source": file_path}

def process_pdf(file_path):
    with open(file_path, 'rb') as file:
        reader = PyPDF2.PdfReader(file)
        content = []
        for page in reader.pages:
            content.append(page.extract_text())
    return "\n".join(content), {"source": file_path}

def process_image(file_path):
    image = Image.open(file_path)
    content = pytesseract.image_to_string(image)
    return content, {"source": file_path, "image_path": file_path}
